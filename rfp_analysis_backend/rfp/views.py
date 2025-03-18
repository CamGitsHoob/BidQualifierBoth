import os
import uuid
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.http import JsonResponse, HttpResponse
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser
from PyPDF2 import PdfReader
from haystack import Document
from haystack.components.preprocessors import DocumentSplitter
from haystack.components.embedders import OpenAIDocumentEmbedder
from haystack.utils import Secret
from pinecone_store import document_store, get_document_store, reset_document_store
from .rfp_analyzer import RFPAnalyzer
from asgiref.sync import async_to_sync
from .rfp_chatbot import RFPChatbot
from rest_framework.response import Response
from rest_framework import status
import json
import asyncio
from pinecone_store import Pinecone
import numpy as np
from openpyxl import Workbook
from datetime import datetime
from pinecone_store import get_session_index_name, pc
from django.conf import settings

# Create a global analyzer instance using our Pinecone document store.
analyzer = RFPAnalyzer(vector_store=document_store)

def extract_text_from_pdf(file_path):
    """
    Extract text from a PDF file.
    """
    try:
        # Check if the file exists at the given path
        if not os.path.exists(file_path):
            # Try to get the absolute path using default_storage
            absolute_path = default_storage.path(file_path)
            if not os.path.exists(absolute_path):
                raise FileNotFoundError(f"File not found at {file_path} or {absolute_path}")
            file_path = absolute_path
        
        with open(file_path, "rb") as f:
            reader = PdfReader(f)
            extracted_text = "\n".join(
                [page.extract_text() for page in reader.pages if page.extract_text()]
            )
            if not extracted_text:
                raise ValueError("No extractable text found in PDF")
            return extracted_text
    except Exception as e:
        print(f"PDF extraction failed: {e}")
        raise

@api_view(["POST"])
@parser_classes([MultiPartParser])
def upload_pdf(request):
    """
    Upload an RFP PDF file, extract text, split it into chunks,
    compute 1536-d OpenAI embeddings, and index them into Pinecone.
    """
    try:
        file = request.FILES.get("file")
        if not file or not file.name.endswith(".pdf"):
            return JsonResponse({"error": "Invalid file"}, status=400)

        # Reset the document store for new upload
        print("Resetting document store")
        global document_store
        document_store = reset_document_store()

        # Generate a unique identifier for this document
        unique_id = str(uuid.uuid4())
        file_path = f"rfp_documents/{unique_id}_{file.name}"
        file_name = default_storage.save(file_path, ContentFile(file.read()))
        print(f"Saved PDF at: {default_storage.path(file_name)}")

        # Extract text from the PDF
        try:
            extracted_text = extract_text_from_pdf(default_storage.path(file_name))
        except Exception as e:
            return JsonResponse({"error": f"Failed to read PDF: {str(e)}"}, status=500)

        # Split the text into document chunks
        splitter = DocumentSplitter(split_by="sentence", split_length=3, split_overlap=1)
        splitter.warm_up()
        docs = [Document(content=extracted_text)]
        split_docs = splitter.run(docs)["documents"]
        print(f"Split into {len(split_docs)} document chunks")

        # Get OpenAI API key
        openai_api_key = os.environ.get("OPENAI_API_KEY")
        if not openai_api_key:
            return JsonResponse(
                {"error": "OPENAI_API_KEY environment variable is not set."},
                status=500,
            )

        # Embed documents
        document_embedder = OpenAIDocumentEmbedder(
            api_key=Secret.from_token(openai_api_key),
            model="text-embedding-ada-002"
        )
        embedding_results = document_embedder.run(split_docs)
        embedded_docs = embedding_results["documents"]

        # Debug: Log embedding dimensions
        for i, doc in enumerate(embedded_docs):
            if doc.embedding:
                print(f"Document {i} embedding dimension: {len(doc.embedding)}")
            else:
                print(f"Document {i} has no embedding.")

        # Index documents
        document_store.write_documents(embedded_docs)
        print("Documents written to Pinecone document store.")

        # Clean up the file after processing
        default_storage.delete(file_name)

        return JsonResponse({
            "success": True,
            "message": "File uploaded and processed successfully"
        })

    except Exception as e:
        print(f"Error in upload_pdf: {str(e)}")
        return JsonResponse({
            "error": f"Upload failed: {str(e)}"
        }, status=500)

@api_view(["POST"])
@parser_classes([MultiPartParser])
def analyze_pdf(request):
    """Process and index the PDF in Pinecone."""
    try:
        print("analyze_pdf view called")
        print("Request data:", request.data)
        print("Request FILES:", request.FILES)
        
        # Generate a session ID if not provided
        session_id = request.data.get('session_id')
        if not session_id:
            session_id = str(uuid.uuid4())
            print(f"Generated new session ID: {session_id}")
        
        # Get the file from the request
        if 'file' not in request.FILES:
            print("No file in request.FILES")
            return JsonResponse({"error": "No file provided"}, status=400)
        
        uploaded_file = request.FILES['file']
        print(f"Received file: {uploaded_file.name}, size: {uploaded_file.size}")
        
        # Save the file temporarily
        file_path = default_storage.save(f"uploads/{uploaded_file.name}", ContentFile(uploaded_file.read()))
        print(f"Saved file to: {file_path}")
        
        # Get the absolute path to the file
        absolute_file_path = default_storage.path(file_path)
        print(f"Absolute file path: {absolute_file_path}")
        
        # Reset the document store for this session
        document_store = reset_document_store(session_id)
        print(f"Reset document store for session: {session_id}")

        # Extract and process the PDF
        extracted_text = extract_text_from_pdf(absolute_file_path)
        print(f"Extracted text length: {len(extracted_text)}")
        
        # Split into chunks and embed
        splitter = DocumentSplitter(split_by="sentence", split_length=3, split_overlap=1)
        # Warm up the splitter before using it
        splitter.warm_up()
        
        docs = [Document(content=extracted_text)]
        split_docs = splitter.run(docs)["documents"]

        # Get OpenAI API key
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            return JsonResponse({"error": "OpenAI API key not found"}, status=500)

        # Embed the documents
        embedder = OpenAIDocumentEmbedder(api_key=Secret.from_token(api_key))
        
        embedded_docs = embedder.run(split_docs)["documents"]

        # Write to Pinecone
        document_store.write_documents(embedded_docs)

        # Clean up the temporary file
        default_storage.delete(file_path)

        return JsonResponse({
            "success": True,
            "message": "Document analyzed and indexed successfully",
            "session_id": session_id
        })

    except Exception as e:
        import traceback
        print(f"Error in analyze_pdf: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            "error": f"Analysis failed: {str(e)}"
        }, status=500)

@api_view(["POST"])
def analyze_rfp(request):
    """Analyze the RFP using the indexed documents."""
    try:
        # Get the session ID
        session_id = request.data.get('session_id')
        
        # Get the document store for this session
        document_store = get_document_store(session_id)
        
        # Create an analyzer with the session-specific document store
        analyzer = RFPAnalyzer(vector_store=document_store)

        
        # Use the analyzer to analyze the indexed documents
        result = async_to_sync(analyzer.analyze_rfp)("", "")
        
        return JsonResponse({
            "success": True,
            "result": result,
            "session_id": session_id
        })

    except Exception as e:
        import traceback
        print(f"Error in analyze_rfp: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            "error": f"Analysis failed: {str(e)}"
        }, status=500)

@api_view(["POST"])
def generate_bid_matrix(request, doc_id):
    """
    Generate a bid matrix based on an RFP document.
    """
    result = async_to_sync(analyzer.generate_bid_matrix)({"doc_id": doc_id})
    return JsonResponse({"matrix": result})

@api_view(["GET"])
def download_matrix(request, doc_id):
    """
    Download the bid matrix as an Excel file.
    """
    import io
    import pandas as pd

    data = {
        "Requirement": [],
        "Priority": [],
        "Complexity": [],
        "Status": [],
        "Assigned To": [],
        "Notes": [],
    }
    df = pd.DataFrame(data)
    excel_buffer = io.BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)

    response = HttpResponse(
        excel_buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename=bid_matrix_{doc_id}.xlsx"
    return response

@api_view(['POST'])
def chat_with_rfp(request):
    """
    Endpoint to chat with all RFP documents in the index
    """
    try:
        print("Chat endpoint called")
        print(f"Request data: {request.data}")
        
        question = request.data.get('question')
        print(f"Extracted question: {question}")
        
        if not question:
            print("No question provided")
            return Response(
                {"error": "Question is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Initialize chatbot without analysis_id
            print("Initializing chatbot")
            chatbot = RFPChatbot()
            
            # Get response
            print("Getting response from chatbot")
            response = chatbot.get_response(question)
            print(f"Chatbot response: {response}")

            if not response:
                raise ValueError("Empty response from chatbot")

            return Response({
                "answer": response,
                "success": True
            })

        except Exception as chat_error:
            print(f"Error in chatbot: {str(chat_error)}")
            raise  # Re-raise to be caught by outer try-except

    except Exception as e:
        import traceback
        print(f"Error in chat_with_rfp: {str(e)}")
        print("Full traceback:")
        print(traceback.format_exc())
        return Response(
            {
                "error": f"Failed to process chat request: {str(e)}",
                "success": False
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(["GET"])
def compare_indexes(request):
    """Compare documents in uswebbid against rfp-index for similarity."""
    try:
        # Initialize Pinecone
        pc = Pinecone(api_key=os.getenv("PINECONE_API_KEY"))
        
        # Get both indexes
        uswebbid = pc.Index("paidmediabids")
        rfp = pc.Index("rfpupload")
        
        # Get vectors from rfp-index
        rfp_response = rfp.query(
            vector=[1.0] * 1536,
            top_k=400,
            include_values=True,
            include_metadata=True,
            namespace="default"
        ).to_dict()

        # Get the first vector from rfp-index results
        if rfp_response.get('matches'):
            rfp_vector = rfp_response['matches'][0]['values']
            
            # Use this vector to query uswebbid
            uswebbid_response = uswebbid.query(
                vector=rfp_vector,
                top_k=400,
                include_values=True,
                include_metadata=True,
                namespace="default"
            ).to_dict()
            
            # Calculate average similarity score
            scores = [match['score'] for match in uswebbid_response.get('matches', [])]
            average_similarity = sum(scores) / len(scores) if scores else 0
            
            return JsonResponse({
                'success': True,
                'similarity_score': round(average_similarity, 4),
                'total_documents_compared': len(scores)
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No vectors found in rfpupload index'
            })

    except Exception as e:
        print(f"Error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@api_view(["POST"])  # Change to POST to receive data
def download_report(request):
    """Download the RFP analysis as an Excel report."""
    try:
        print("Starting report generation...")
        
        # Get the RFP data from the request
        rfp_data = request.data.get('rfpData', {})
        print(f"Received RFP data: {rfp_data}")  # Debug log

        wb = Workbook()
        ws = wb.active
        ws.title = "RFP Analysis Report"

        # Add title and date
        ws['A1'] = "RFP Analysis Report"
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        
        current_row = 4

        def add_section(title, data):
            nonlocal current_row
            try:
                # Add section title
                ws.cell(row=current_row, column=1, value=title)
                ws.merge_cells(f'A{current_row}:F{current_row}')
                current_row += 1

                # Add data
                if isinstance(data, dict):
                    for key, value in data.items():
                        ws.cell(row=current_row, column=1, value=str(key).replace('_', ' ').title())
                        ws.cell(row=current_row, column=2, value=str(value))
                        current_row += 1
                
                current_row += 1
            except Exception as e:
                print(f"Error in add_section for {title}: {str(e)}")
                raise

        # Map the sections to the RFP data structure
        sections = {
            "Strategic Summary": rfp_data.get('strategic_summary', {}),
            "Introduction/Background": rfp_data.get('introduction/background', {}),
            "Bid Summary": rfp_data.get('bid_summary', {}),
            "Key Dates": rfp_data.get('key_dates', {}),
            "Work Portfolio": rfp_data.get('work_portfolio', {}),
            "Requirements": rfp_data.get('requirements', {}),
            "Website Details": rfp_data.get('website_details', {}),
            "Commercials": rfp_data.get('commercials', {}),
            "Flags": rfp_data.get('flags', {})
        }

        for section_title, section_data in sections.items():
            print(f"Processing section: {section_title}")
            add_section(section_title, section_data)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=rfp_analysis_report.xlsx'

        wb.save(response)
        print("Report generated successfully")
        return response

    except Exception as e:
        print(f"Error generating report: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@api_view(["POST"])
def cleanup_session(request):
    """Clean up a session's resources."""
    try:
        session_id = request.data.get('session_id')
        if not session_id:
            return JsonResponse({"error": "No session ID provided"}, status=400)
        
        # Get the index name for this session
        index_name = get_session_index_name(session_id)
        
        # SAFETY CHECK: Only delete if it's a session index
        # This prevents deletion of static indexes
        if index_name.startswith(f"{index_name_base}-") and index_name != index_name_base:
            if index_name in pc.list_indexes().names():
                # Check if the index is protected
                if index_name in getattr(settings, 'PROTECTED_INDEXES', []):
                    return JsonResponse({
                        "error": "Cannot delete protected index",
                        "message": f"Index {index_name} is protected"
                    }, status=403)
                else:
                    pc.delete_index(index_name)
                    return JsonResponse({
                        "success": True,
                        "message": f"Session {session_id} cleaned up successfully"
                    })
            else:
                return JsonResponse({
                    "success": True,
                    "message": f"No index found for session {session_id}"
                })
        else:
            return JsonResponse({
                "error": "Cannot delete non-session index",
                "message": f"Index {index_name} appears to be a static index"
            }, status=403)
            
    except Exception as e:
        return JsonResponse({
            "error": f"Failed to clean up session: {str(e)}"
        }, status=500)